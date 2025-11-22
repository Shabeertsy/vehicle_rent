from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum, Count, F
from django.db.models.functions import TruncMonth
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.models import User
from django.contrib.auth.models import User
from django.views.decorators.http import require_POST
from decimal import Decimal
from .models import Vehicle, Rental, Expense, UserProfile, TakenAmount
from .forms import VehicleForm, RentalForm, ExpenseForm, UserCreateForm, UserEditForm
import pandas as pd
from datetime import datetime, date

def dashboard(request):
    # Overall Analytics
    total_income = Rental.objects.aggregate(Sum('total_amount_received'))['total_amount_received__sum'] or 0
    total_expense = Expense.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    profit = total_income - total_expense
    active_vehicles_count = Vehicle.objects.count()

    # Monthly Data for Graph and Table
    # Group rentals by month
    rentals_by_month = Rental.objects.annotate(month=TruncMonth('date_out')).values('month').annotate(income=Sum('total_amount_received')).order_by('month')

    # Group expenses by month
    expenses_by_month = Expense.objects.annotate(month=TruncMonth('date')).values('month').annotate(expense=Sum('amount')).order_by('month')

    # Merge data
    monthly_data = {}

    for r in rentals_by_month:
        month = r['month'].strftime('%B %Y')
        if month not in monthly_data:
            monthly_data[month] = {'month': month, 'income': 0, 'expense': 0, 'profit': 0}
        monthly_data[month]['income'] += float(r['income'])

    for e in expenses_by_month:
        month = e['month'].strftime('%B %Y')
        if month not in monthly_data:
            monthly_data[month] = {'month': month, 'income': 0, 'expense': 0, 'profit': 0}
        monthly_data[month]['expense'] += float(e['expense'])

    # Calculate profit for each month
    final_monthly_data = []
    for month, data in monthly_data.items():
        data['profit'] = data['income'] - data['expense']
        final_monthly_data.append(data)

    # Sort by date (parsing the month string back to date for sorting)
    final_monthly_data.sort(key=lambda x: datetime.strptime(x['month'], '%B %Y'))

    context = {
        'total_income': total_income,
        'total_expense': total_expense,
        'profit': profit,
        'active_vehicles_count': active_vehicles_count,
        'monthly_data': final_monthly_data,
    }
    return render(request, 'dashboard.html', context)

# Vehicle Views
def vehicle_list(request):
    vehicles = Vehicle.objects.all()
    return render(request, 'vehicle_list.html', {'vehicles': vehicles})

def vehicle_detail(request, pk):
    vehicle = get_object_or_404(Vehicle, pk=pk)
    rentals = vehicle.rentals.all().order_by('-date_out')
    expenses = vehicle.expenses.all().order_by('-date')

    total_revenue = rentals.aggregate(Sum('total_amount_received'))['total_amount_received__sum'] or 0
    total_expense = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    profit = total_revenue - total_expense

    # Monthly breakdown for this vehicle (income, expense, profit per month) including all months
    # Rentals grouped by month
    rentals_by_month = rentals.annotate(month=TruncMonth('date_out')).values('month').annotate(income=Sum('total_amount_received')).order_by('month')
    # Expenses grouped by month
    expenses_by_month = expenses.annotate(month=TruncMonth('date')).values('month').annotate(expense=Sum('amount')).order_by('month')

    # Figure out the month range present in rentals/expenses
    min_month = None
    max_month = None

    for queryset in [rentals_by_month, expenses_by_month]:
        for d in queryset:
            if d['month']:
                if not min_month or d['month'] < min_month:
                    min_month = d['month']
                if not max_month or d['month'] > max_month:
                    max_month = d['month']

    if not min_month or not max_month:
        today = date.today().replace(day=1)
        min_month = max_month = today

    # Build a list of all months from min_month to max_month
    current = min_month.replace(day=1)
    months = []
    while current <= max_month:
        months.append(current)
        # Move to next month
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)

    # Prepare lookups
    rental_map = {r['month']: float(r['income'] or 0) for r in rentals_by_month if r['month']}
    expense_map = {e['month']: float(e['expense'] or 0) for e in expenses_by_month if e['month']}

    monthly_data = []
    for month_dt in months:
        income = rental_map.get(month_dt, 0)
        expense = expense_map.get(month_dt, 0)
        profit_val = income - expense
        monthly_data.append({
            'month': month_dt,  # Pass the date object directly
            'income': income,
            'expense': expense,
            'profit': profit_val
        })

    # Pass months (as list of date objects) to the front end, so template can access all months explicitly.

    # Check EMI status
    emi_warning = False
    emi_due_date = None
    emi_is_paid = False
    days_until_emi = None

    try:
        emi_config = vehicle.emi
        if emi_config.is_active:
            today = date.today()

            # Calculate due date for this month
            try:
                due_date = date(today.year, today.month, emi_config.due_day)
            except ValueError:
                # Handle months with fewer days (e.g., Feb 30 -> Feb 28/29)
                import calendar
                last_day = calendar.monthrange(today.year, today.month)[1]
                due_date = date(today.year, today.month, last_day)

            emi_due_date = due_date

            # Check if paid using EMIPayment model
            from .models import EMIPayment
            emi_is_paid = EMIPayment.objects.filter(
                vehicle=vehicle,
                date__year=today.year,
                date__month=today.month
            ).exists()

            if not emi_is_paid:
                days_until_emi = (due_date - today).days
                if days_until_emi <= emi_config.warning_days:
                    emi_warning = True
    except Exception:
        # No EMI configured for this vehicle or other error
        emi_config = None

    # Get EMI history
    from .models import EMIPayment
    emi_payments = EMIPayment.objects.filter(vehicle=vehicle).order_by('-date')

    context = {
        'vehicle': vehicle,
        'rentals': rentals,
        'expenses': expenses,
        'emi_payments': emi_payments,  # Add this line
        'total_revenue': total_revenue,
        'total_expense': total_expense,
        'profit': profit,
        'monthly_data': monthly_data,
        'all_months': months,   # <<<This is the additional context variable
        'emi_warning': emi_warning,
        'emi_due_date': emi_due_date,
        'emi_is_paid': emi_is_paid,
        'emi_config': emi_config,
        'days_until_emi': days_until_emi,
    }
    return render(request, 'vehicle_detail.html', context)

def vehicle_create(request):
    if request.method == 'POST':
        form = VehicleForm(request.POST, request.FILES)
        if form.is_valid():
            vehicle = form.save()

            # Handle partners manually from POST data
            partner_ids = request.POST.getlist('partners')
            if partner_ids:
                partners = User.objects.filter(pk__in=partner_ids, is_active=True)
                vehicle.partners.set(partners)
            else:
                vehicle.partners.clear()

            messages.success(request, 'Vehicle created successfully.')
            return redirect('vehicle_list')
    else:
        form = VehicleForm()

    partners = User.objects.filter(is_active=True).order_by('username')
    return render(request, 'form.html', {'form': form, 'title': 'Add Vehicle', 'partners': partners})

def vehicle_edit(request, pk):
    vehicle = get_object_or_404(Vehicle, pk=pk)
    if request.method == 'POST':
        form = VehicleForm(request.POST, request.FILES, instance=vehicle)
        if form.is_valid():
            vehicle = form.save()

            # Handle partners manually from POST data
            partner_ids = request.POST.getlist('partners')
            if partner_ids:
                partners = User.objects.filter(pk__in=partner_ids, is_active=True)
                vehicle.partners.set(partners)
            else:
                vehicle.partners.clear()

            messages.success(request, 'Vehicle updated successfully.')
            return redirect('vehicle_detail', pk=pk)
    else:
        form = VehicleForm(instance=vehicle)

    partners = User.objects.filter(is_active=True).order_by('username')
    selected_partners = vehicle.partners.all()
    return render(request, 'form.html', {
        'form': form,
        'title': 'Edit Vehicle',
        'partners': partners,
        'selected_partners': selected_partners,
        'vehicle': vehicle
    })

def vehicle_delete(request, pk):
    vehicle = get_object_or_404(Vehicle, pk=pk)
    if request.method == 'POST':
        vehicle.delete()
        messages.success(request, 'Vehicle deleted successfully.')
        return redirect('vehicle_list')
    return render(request, 'confirm_delete.html', {'object': vehicle})

# Rental Views
def rental_create(request, vehicle_id):
    vehicle = get_object_or_404(Vehicle, pk=vehicle_id)
    if request.method == 'POST':
        form = RentalForm(request.POST)
        if form.is_valid():
            rental = form.save(commit=False)
            rental.vehicle = vehicle
            rental.save()

            # Send email notification to partners
            from .notifications import send_partner_notification
            send_partner_notification(vehicle, 'rental', {
                'customer_name': rental.customer_name,
                'date_out': rental.date_out,
                'destination': rental.destination or 'N/A',
                'days_of_rent': rental.days_of_rent,
                'total_amount_received': rental.total_amount_received,
            })

            messages.success(request, 'Rental added successfully.')
            return redirect('vehicle_detail', pk=vehicle_id)
    else:
        initial_data = {'rent_per_day': vehicle.price_per_day}
        form = RentalForm(initial=initial_data)
    return render(request, 'form.html', {'form': form, 'title': f'Add Rental for {vehicle.name}'})

def rental_edit(request, pk):
    rental = get_object_or_404(Rental, pk=pk)
    if request.method == 'POST':
        form = RentalForm(request.POST, instance=rental)
        if form.is_valid():
            form.save()
            messages.success(request, 'Rental updated successfully.')
            return redirect('vehicle_detail', pk=rental.vehicle.id)
    else:
        form = RentalForm(instance=rental)
    return render(request, 'form.html', {'form': form, 'title': 'Edit Rental'})

def rental_delete(request, pk):
    rental = get_object_or_404(Rental, pk=pk)
    vehicle_id = rental.vehicle.id
    if request.method == 'POST':
        rental.delete()
        messages.success(request, 'Rental deleted successfully.')
        return redirect('vehicle_detail', pk=vehicle_id)
    return render(request, 'confirm_delete.html', {'object': rental})

# Expense Views
def expense_create(request, vehicle_id):
    vehicle = get_object_or_404(Vehicle, pk=vehicle_id)
    if request.method == 'POST':
        form = ExpenseForm(request.POST)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.vehicle = vehicle
            expense.save()

            # Send email notification to partners
            from .notifications import send_partner_notification
            send_partner_notification(vehicle, 'expense', {
                'date': expense.date,
                'particulars': expense.particulars,
                'place': expense.place or 'N/A',
                'amount': expense.amount,
            })

            messages.success(request, 'Expense added successfully.')
            return redirect('vehicle_detail', pk=vehicle_id)
    else:
        form = ExpenseForm()
    return render(request, 'form.html', {'form': form, 'title': f'Add Expense for {vehicle.name}'})

def expense_edit(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, 'Expense updated successfully.')
            return redirect('vehicle_detail', pk=expense.vehicle.id)
    else:
        form = ExpenseForm(instance=expense)
    return render(request, 'form.html', {'form': form, 'title': 'Edit Expense'})

def expense_delete(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    vehicle_id = expense.vehicle.id
    if request.method == 'POST':
        expense.delete()
        messages.success(request, 'Expense deleted successfully.')
        return redirect('vehicle_detail', pk=vehicle_id)
    return render(request, 'confirm_delete.html', {'object': expense})

# Excel Import
def import_data(request):
    vehicles = Vehicle.objects.all()
    import_errors = []
    selected_vehicle_id = None

    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        vehicle_id = request.POST.get('vehicle_id')
        selected_vehicle_id = vehicle_id

        if not vehicle_id:
            messages.error(request, 'Please select a vehicle.')
            return render(request, 'import.html', {'vehicles': vehicles, 'import_errors': import_errors, 'selected_vehicle_id': selected_vehicle_id})

        vehicle = get_object_or_404(Vehicle, pk=vehicle_id)

        try:
            # Try to read the Excel file from the beginning each time
            # For seekability, use InMemoryUploadedFile, so read into BytesIO if needed
            from io import BytesIO
            excel_data = excel_file.read()
            # Reset pointer for every read_excel
            def get_excel_stream():
                return BytesIO(excel_data)

            header_row_index = 0
            found_header = False

            # Step 1: Find the header row
            df_temp = pd.read_excel(get_excel_stream(), header=None, nrows=20)
            for i, row in df_temp.iterrows():
                row_values = [str(val).strip().upper() for val in row.values]
                if 'DATE OUT' in row_values and 'CUSTOMER' in row_values:
                    header_row_index = i
                    found_header = True
                    break

            if not found_header:
                messages.warning(request, "Could not find 'DATE OUT' header row. Please check file format.")
                return render(request, 'import.html', {'vehicles': vehicles, 'import_errors': ['Header row not found (DATE OUT, CUSTOMER)'], 'selected_vehicle_id': selected_vehicle_id})

            # Step 2: Read the full file with the correct header
            df = pd.read_excel(get_excel_stream(), header=header_row_index)

            # Normalize column names
            df.columns = [str(col).strip().upper() for col in df.columns]

            rental_success_count = 0
            expense_success_count = 0

            for index, row in df.iterrows():
                # --- Process Rental ---
                try:
                    date_out_raw = row.get('DATE OUT')
                    customer_raw = row.get('CUSTOMER')

                    if pd.notna(date_out_raw) and pd.notna(customer_raw):
                        # Parse Date Out
                        date_out = None
                        try:
                            if isinstance(date_out_raw, str):
                                date_out = pd.to_datetime(date_out_raw, dayfirst=True, errors='coerce').date()
                            elif hasattr(date_out_raw, 'date'):
                                date_out = date_out_raw.date()
                            elif isinstance(date_out_raw, pd.Timestamp):
                                date_out = date_out_raw.date()
                        except Exception as e:
                            date_out = None

                        if date_out:
                            rental = Rental(
                                vehicle=vehicle,
                                date_out=date_out,
                                customer_name=str(customer_raw).strip(),
                                destination=row.get('DESTINATION'),
                                customer_id=row.get('CUSTOMER ID')
                            )
                            # Contact No (sometimes ; at end)
                            rental.contact_no = row.get('CONTACT NO;') or row.get('CONTACT NO')
                            rental.care_of = row.get('C/O')

                            # Time Out
                            time_out_raw = row.get('TIME OUT')
                            rental.time_out = None
                            if pd.notna(time_out_raw):
                                try:
                                    if hasattr(time_out_raw, 'time'):
                                        rental.time_out = time_out_raw
                                    elif isinstance(time_out_raw, str):
                                        parsed_time = pd.to_datetime(time_out_raw, errors='coerce')
                                        if not pd.isna(parsed_time):
                                            rental.time_out = parsed_time.time()
                                except: pass

                            # Date In
                            date_in_raw = row.get('DATE IN')
                            rental.date_in = None
                            if pd.notna(date_in_raw):
                                try:
                                    if hasattr(date_in_raw, 'date'):
                                        rental.date_in = date_in_raw.date()
                                    elif isinstance(date_in_raw, str):
                                        rental.date_in = pd.to_datetime(date_in_raw, dayfirst=True, errors='coerce').date()
                                except: pass

                            # Time In
                            time_in_raw = row.get('TIME IN')
                            rental.time_in = None
                            if pd.notna(time_in_raw):
                                try:
                                    if hasattr(time_in_raw, 'time'):
                                        rental.time_in = time_in_raw
                                    elif isinstance(time_in_raw, str):
                                        parsed_time = pd.to_datetime(time_in_raw, errors='coerce')
                                        if not pd.isna(parsed_time):
                                            rental.time_in = parsed_time.time()
                                except: pass

                            # Numeric fields
                            rental.days_of_rent = pd.to_numeric(row.get('DAYS OF RENT'), errors='coerce')
                            if pd.isna(rental.days_of_rent): rental.days_of_rent = 0
                            rental.rent_per_day = pd.to_numeric(row.get('RENT/DAY'), errors='coerce')
                            if pd.isna(rental.rent_per_day): rental.rent_per_day = 0
                            adv_amt = None
                            if 'ADV; AMOUNT' in df.columns:
                                adv_amt = row.get('ADV; AMOUNT')
                            if adv_amt is None and 'ADV AMOUNT' in df.columns:
                                adv_amt = row.get('ADV AMOUNT')
                            rental.advance_amount = pd.to_numeric(adv_amt, errors='coerce')
                            if pd.isna(rental.advance_amount): rental.advance_amount = 0

                            rental.starting_km = pd.to_numeric(row.get('STARTING KM'), errors='coerce')
                            if pd.isna(rental.starting_km): rental.starting_km = None
                            rental.ending_km = pd.to_numeric(row.get('ENDING KM'), errors='coerce')
                            if pd.isna(rental.ending_km): rental.ending_km = None
                            rental.total_amount_received = pd.to_numeric(row.get('TOTAL AMOUNT RECEIVED'), errors='coerce')
                            if pd.isna(rental.total_amount_received): rental.total_amount_received = 0

                            rental.save()
                            rental_success_count += 1

                except Exception as e:
                    import_errors.append(f"Rental Row {index+1}: {e}")

                # --- Process Expense ---
                try:
                    particulars = row.get('PARTICULARS')
                    amount = row.get('AMOUNT')

                    if pd.notna(particulars) and pd.notna(amount):
                        # Expense Date
                        date_val_raw = row.get('DATE')
                        expense_date = timezone.now().date()
                        if pd.notna(date_val_raw):
                            try:
                                if hasattr(date_val_raw, 'date'):
                                    expense_date = date_val_raw.date()
                                elif isinstance(date_val_raw, str):
                                    date_parsed = pd.to_datetime(date_val_raw, dayfirst=True, errors='coerce')
                                    if not pd.isna(date_parsed):
                                        expense_date = date_parsed.date()
                            except: pass

                        # Handle C/O for expense rows (C/O.1 or fallback if no CUSTOMER)
                        care_of_expense = None
                        if 'C/O.1' in df.columns:
                            care_of_expense = row.get('C/O.1')
                        if not care_of_expense and 'C/O' in df.columns and pd.isna(row.get('CUSTOMER')):
                            care_of_expense = row.get('C/O')

                        expense_amount = pd.to_numeric(amount, errors='coerce')
                        if pd.isna(expense_amount): expense_amount = 0

                        expense = Expense(
                            vehicle=vehicle,
                            date=expense_date,
                            particulars=particulars,
                            place=row.get('PLACE'),
                            care_of=care_of_expense,
                            amount=expense_amount
                        )
                        expense.save()
                        expense_success_count += 1

                except Exception as e:
                    import_errors.append(f"Expense Row {index+1}: {e}")

            msg = []
            if rental_success_count > 0:
                msg.append(f'{rental_success_count} rentals')
            if expense_success_count > 0:
                msg.append(f'{expense_success_count} expenses')
            if msg:
                messages.success(request, f"Successfully imported: {', '.join(msg)}.")
            else:
                if not found_header:
                    messages.warning(request, "Could not find 'DATE OUT' header row. Please check file format.")
                else:
                    messages.warning(request, 'No records were imported. Please check the file format and ensure data exists.')
            if import_errors:
                messages.error(request, f"Some rows were skipped due to errors. See details below.")
                return render(request, 'import.html', {'vehicles': vehicles, 'import_errors': import_errors, 'selected_vehicle_id': selected_vehicle_id})
        except Exception as e:
            import_errors.append(str(e))
            messages.error(request, f'Error importing file: {str(e)}')
            return render(request, 'import.html', {'vehicles': vehicles, 'import_errors': import_errors, 'selected_vehicle_id': selected_vehicle_id})

        return redirect('vehicle_detail', pk=vehicle_id)

    if request.method == 'GET':
        selected_vehicle_id = request.GET.get('vehicle_id')

    return render(request, 'import.html', {'vehicles': vehicles, 'import_errors': import_errors, 'selected_vehicle_id': selected_vehicle_id})


# User Management Views
def user_list(request):
    """Display list of all users"""
    users = User.objects.all().order_by('-date_joined')
    active_count = users.filter(is_active=True).count()
    inactive_count = users.filter(is_active=False).count()

    context = {
        'users': users,
        'active_count': active_count,
        'inactive_count': inactive_count,
    }
    return render(request, 'user_list.html', context)

def user_detail(request, pk):
    user = get_object_or_404(User, pk=pk)

    # Get filter year, default to current year
    current_year = datetime.now().year
    selected_year = request.GET.get('year', current_year)
    try:
        selected_year = int(selected_year)
    except ValueError:
        selected_year = current_year

    # Filter rentals and expenses by user and year (for context, though not used in monthly breakdown)
    rentals = Rental.objects.filter(user=user, date_out__year=selected_year).order_by('-date_out')
    expenses = Expense.objects.filter(user=user, date__year=selected_year).order_by('-date')

    # Calculate totals for the selected year (based on profit sharing)
    total_income = 0
    total_expense = 0
    profit = 0

    # Monthly breakdown data structures (using month numbers 1-12 as keys)
    monthly_shares = {}  # For income/expense from vehicles
    monthly_taken = {}   # For taken amounts

    # 1. Calculate monthly share from vehicles
    # Get all vehicles where user is a partner
    vehicles = user.vehicles.all()

    for vehicle in vehicles:
        num_partners = vehicle.partners.count()
        if num_partners > 0:
            # Get monthly stats for this vehicle for the selected year
            v_rentals = vehicle.rentals.filter(date_out__year=selected_year).annotate(month=TruncMonth('date_out')).values('month').annotate(income=Sum('total_amount_received'))
            v_expenses = vehicle.expenses.filter(date__year=selected_year).annotate(month=TruncMonth('date')).values('month').annotate(expense=Sum('amount'))

            for r in v_rentals:
                m_idx = r['month'].month
                if m_idx not in monthly_shares:
                    monthly_shares[m_idx] = {'income': 0, 'expense': 0}
                monthly_shares[m_idx]['income'] += float(r['income']) / num_partners

            for e in v_expenses:
                m_idx = e['month'].month
                if m_idx not in monthly_shares:
                    monthly_shares[m_idx] = {'income': 0, 'expense': 0}
                monthly_shares[m_idx]['expense'] += float(e['expense']) / num_partners

    # 2. Calculate monthly taken amounts
    taken_amounts = TakenAmount.objects.filter(user=user, date__year=selected_year).order_by('-date')
    taken_by_month = taken_amounts.annotate(month=TruncMonth('date')).values('month').annotate(amount=Sum('amount')).order_by('month')

    for t in taken_by_month:
        m_idx = t['month'].month
        monthly_taken[m_idx] = float(t['amount'])

    # 3. Merge into final monthly data
    final_monthly_data = []

    # Helper for month names
    import calendar

    for m_idx in range(1, 13):
        month_name = calendar.month_name[m_idx]
        data = {
            'month': month_name,
            'income': 0,
            'expense': 0,
            'profit': 0,
            'taken': 0
        }

        # Add share data
        if m_idx in monthly_shares:
            data['income'] = monthly_shares[m_idx]['income']
            data['expense'] = monthly_shares[m_idx]['expense']
            data['profit'] = data['income'] - data['expense']

        # Add taken data
        if m_idx in monthly_taken:
            data['taken'] = monthly_taken[m_idx]

        # Only add if there's data
        if data['income'] != 0 or data['expense'] != 0 or data['taken'] != 0:
            final_monthly_data.append(data)

    # Calculate vehicle-based profit shares and taken amounts (All time or selected year? Usually all time for balance)
    # But for the table we might want to show selected year stats?
    # The balance is usually a running balance.

    vehicle_data = []
    total_profit_share = Decimal('0')
    total_taken = Decimal('0')

    for vehicle in vehicles:
        # Get total profit for the vehicle (All time)
        v_rentals_total = vehicle.rentals.aggregate(Sum('total_amount_received'))['total_amount_received__sum'] or 0
        v_expenses_total = vehicle.expenses.aggregate(Sum('amount'))['amount__sum'] or 0
        v_profit = Decimal(str(v_rentals_total)) - Decimal(str(v_expenses_total))

        # Calculate share based on number of partners
        num_partners = vehicle.partners.count()
        user_share = Decimal('0')
        if num_partners > 0:
            user_share = v_profit / num_partners

        # Get taken amount for this vehicle by this user (All time)
        vehicle_taken = TakenAmount.objects.filter(
            user=user,
            vehicle=vehicle
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

        # Calculate balance for this vehicle
        vehicle_balance = user_share - Decimal(str(vehicle_taken))

        vehicle_data.append({
            'vehicle': vehicle,
            'total_profit': v_profit,
            'num_partners': num_partners,
            'user_share': user_share,
            'taken_amount': vehicle_taken,
            'balance': vehicle_balance
        })

        total_profit_share += user_share
        total_taken += Decimal(str(vehicle_taken))

    remaining_balance = total_profit_share - total_taken

    # Calculate totals from final_monthly_data (for the selected year)
    for data in final_monthly_data:
        total_income += data['income']
        total_expense += data['expense']

    profit = total_income - total_expense

    # Get available years for filter
    rental_years = Rental.objects.filter(user=user).dates('date_out', 'year')
    expense_years = Expense.objects.filter(user=user).dates('date', 'year')
    taken_years = TakenAmount.objects.filter(user=user).dates('date', 'year')
    # Also check vehicle rentals where user is partner
    vehicle_rental_years = Rental.objects.filter(vehicle__partners=user).dates('date_out', 'year')

    available_years = sorted(list(set(
        [d.year for d in rental_years] +
        [d.year for d in expense_years] +
        [d.year for d in taken_years] +
        [d.year for d in vehicle_rental_years] +
        [current_year]
    )), reverse=True)

    context = {
        'user_obj': user,
        'rentals': rentals,
        'expenses': expenses,
        'total_income': total_income,
        'total_expense': total_expense,
        'profit': profit,
        'taken_amount': total_taken, # This is all-time taken amount
        'remaining_balance': remaining_balance,
        'monthly_data': final_monthly_data,
        'selected_year': selected_year,
        'available_years': available_years,
        'vehicle_data': vehicle_data,
    }
    return render(request, 'user_detail.html', context)

def user_create(request):
    """Create a new user"""
    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'User created successfully.')
            return redirect('user_list')
    else:
        form = UserCreateForm()
    return render(request, 'form.html', {'form': form, 'title': 'Add User'})

def user_edit(request, pk):
    """Edit an existing user"""
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'User updated successfully.')
            return redirect('user_list')
    else:
        form = UserEditForm(instance=user)
    return render(request, 'form.html', {'form': form, 'title': f'Edit User: {user.username}'})

def user_delete(request, pk):
    """Delete a user"""
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        username = user.username
        user.delete()
        messages.success(request, f'User "{username}" deleted successfully.')
        return redirect('user_list')
    return render(request, 'confirm_delete.html', {'object': user, 'object_type': 'user'})

@require_POST
def update_taken_amount(request, pk):
    """Update the taken amount for a user from a specific vehicle"""
    user = get_object_or_404(User, pk=pk)
    amount = request.POST.get('amount')
    vehicle_id = request.POST.get('vehicle_id')

    if amount and vehicle_id:
        try:
            amount = Decimal(str(amount))
            vehicle = get_object_or_404(Vehicle, pk=vehicle_id)

            # Verify user is a partner of this vehicle
            if not vehicle.partners.filter(pk=user.pk).exists():
                messages.error(request, 'You are not a partner of this vehicle.')
                return redirect('user_detail', pk=pk)

            # Create a record of this transaction
            TakenAmount.objects.create(
                user=user,
                vehicle=vehicle,
                amount=amount,
                date=timezone.now().date()
            )

            messages.success(request, f'Successfully recorded ₹{amount} taken from {vehicle.name}.')
        except (ValueError, TypeError, Exception) as e:
            messages.error(request, f'Error updating amount: {str(e)}')
    else:
        messages.error(request, 'Amount and vehicle are required.')

    return redirect('user_detail', pk=pk)


# Vehicle Partners API
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods

@require_http_methods(["GET"])
def vehicle_partners_get(request, pk):
    """Get all partners and selected partners for a vehicle"""
    vehicle = get_object_or_404(Vehicle, pk=pk)
    all_partners = User.objects.filter(is_active=True).order_by('username')
    selected_partners = vehicle.partners.all()

    return JsonResponse({
        'all_partners': [{'id': p.id, 'username': p.username} for p in all_partners],
        'selected_partners': [p.id for p in selected_partners]
    })

@require_POST
def vehicle_partners_update(request, pk):
    """Update partners for a vehicle"""
    vehicle = get_object_or_404(Vehicle, pk=pk)
    partner_ids = request.POST.getlist('partners')

    if partner_ids:
        partners = User.objects.filter(pk__in=partner_ids, is_active=True)
        vehicle.partners.set(partners)
    else:
        vehicle.partners.clear()

    return JsonResponse({'success': True})

@require_POST
def pay_emi(request, pk):
    """Create an EMI expense for a vehicle"""
    vehicle = get_object_or_404(Vehicle, pk=pk)

    try:
        emi_config = vehicle.emi
        if not emi_config.is_active or emi_config.amount <= 0:
            messages.error(request, 'No active EMI configured for this vehicle.')
            return redirect('vehicle_detail', pk=pk)
    except:
        messages.error(request, 'No EMI configured for this vehicle.')
        return redirect('vehicle_detail', pk=pk)

    today = date.today()

    # Check if EMI for this month has already been paid
    from .models import EMIPayment
    existing_emi = EMIPayment.objects.filter(
        vehicle=vehicle,
        month_paid_for__year=today.year,
        month_paid_for__month=today.month
    ).exists()

    if existing_emi:
        messages.warning(request, 'EMI for this month has already been paid.')
        return redirect('vehicle_detail', pk=pk)

    # Create EMI payment
    emi_payment = EMIPayment.objects.create(
        vehicle=vehicle,
        date=today,
        month_paid_for=today.replace(day=1),
        amount=emi_config.amount,
        remarks=f'EMI Payment for {today.strftime("%B %Y")}'
    )

    # Send email notification to partners
    from .notifications import send_partner_notification
    send_partner_notification(vehicle, 'emi_payment', {
        'amount': emi_config.amount,
        'date': today,
        'month': today.strftime("%B %Y"),
    })

    messages.success(request, f'EMI of ₹{emi_config.amount} has been recorded successfully.')
    return redirect('vehicle_detail', pk=pk)

@require_POST
def delete_emi(request, pk):
    """Delete an EMI payment"""
    from .models import EMIPayment
    emi_payment = get_object_or_404(EMIPayment, pk=pk)
    vehicle_id = emi_payment.vehicle.id
    emi_payment.delete()
    messages.success(request, 'EMI payment deleted successfully.')
    return redirect('vehicle_detail', pk=vehicle_id)

@require_POST
def update_emi(request, pk):
    """Update or create EMI configuration for a vehicle"""
    vehicle = get_object_or_404(Vehicle, pk=pk)

    amount = request.POST.get('emi_amount')
    due_day = request.POST.get('emi_due_day')
    warning_days = request.POST.get('emi_warning_days')
    is_active = request.POST.get('emi_is_active') == 'on'

    if not amount or not due_day or not warning_days:
        messages.error(request, 'All EMI fields are required.')
        return redirect('vehicle_detail', pk=pk)

    try:
        from .models import EMI
        amount = Decimal(str(amount))
        due_day = int(due_day)
        warning_days = int(warning_days)

        # Validate due_day
        if due_day < 1 or due_day > 31:
            messages.error(request, 'Due day must be between 1 and 31.')
            return redirect('vehicle_detail', pk=pk)

        # Validate warning_days
        if warning_days < 0 or warning_days > 30:
            messages.error(request, 'Warning days must be between 0 and 30.')
            return redirect('vehicle_detail', pk=pk)

        # Update or create EMI
        emi, created = EMI.objects.update_or_create(
            vehicle=vehicle,
            defaults={
                'amount': amount,
                'due_day': due_day,
                'warning_days': warning_days,
                'is_active': is_active
            }
        )

        action = 'created' if created else 'updated'
        messages.success(request, f'EMI configuration {action} successfully.')
    except (ValueError, TypeError, Exception) as e:
        messages.error(request, f'Error updating EMI: {str(e)}')

    return redirect('vehicle_detail', pk=pk)


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def vehicle_export_excel(request, pk):
    """Export vehicle data to Excel based on month/year filter - matching exact style"""
    vehicle = get_object_or_404(Vehicle, pk=pk)
    month = request.GET.get('month', 'all')
    year = request.GET.get('year', str(datetime.now().year))

    # Create workbook
    wb = Workbook()

    # Filter data based on month/year
    if month == 'all':
        rentals = vehicle.rentals.all().order_by('date_out')
        expenses = vehicle.expenses.all().order_by('date')
        filename = f"{vehicle.name}_{year}_All_Data.xlsx"
        period_text = f"All Data - {year}"
    else:
        rentals = vehicle.rentals.filter(
            date_out__year=year,
            date_out__month=month
        ).order_by('date_out')
        expenses = vehicle.expenses.filter(
            date__year=year,
            date__month=month
        ).order_by('date')
        month_name = datetime(int(year), int(month), 1).strftime('%B')
        filename = f"{vehicle.name}_{month_name}_{year}.xlsx"
        period_text = f"{month_name} {year}"

    # Styles
    title_font = Font(bold=True, size=14, color="FF0000")  # Red
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    header_font = Font(bold=True, size=10)
    border_thin = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # ===== RENTAL SHEET =====
    ws = wb.active
    ws.title = "Rental History"

    # Calculate totals first
    total_revenue = sum(r.total_amount_received for r in rentals)
    total_expense = sum(e.amount for e in expenses)
    profit = total_revenue - total_expense

    # Title row - Vehicle name and registration
    title_text = f"{vehicle.name.upper()}  {vehicle.registration_number.upper()}"
    ws.merge_cells('A1:P1')
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = title_font
    title_cell.alignment = center_align

    # Financial Summary Section (rows 2-4)
    ws.cell(row=2, column=1, value="Revenue:").font = Font(bold=True, size=11)
    cell_rev = ws.cell(row=2, column=2, value=float(total_revenue))
    cell_rev.font = Font(bold=True, size=11, color="008000")  # Green
    cell_rev.alignment = right_align

    ws.cell(row=3, column=1, value="Expenses:").font = Font(bold=True, size=11)
    cell_exp = ws.cell(row=3, column=2, value=float(total_expense))
    cell_exp.font = Font(bold=True, size=11, color="FF0000")  # Red
    cell_exp.alignment = right_align

    ws.cell(row=4, column=1, value="Profit:").font = Font(bold=True, size=11)
    cell_profit = ws.cell(row=4, column=2, value=float(profit))
    cell_profit.font = Font(bold=True, size=11, color="008000" if profit >= 0 else "FF0000")
    cell_profit.alignment = right_align

    # Vehicle delivery date section (rows 2-5, columns M-O)
    ws.cell(row=2, column=13, value="ADV.").font = Font(bold=True, size=9)
    ws.cell(row=2, column=14, value=vehicle.registration_number).alignment = center_align
    ws.cell(row=3, column=13, value="ADNAN").font = Font(bold=True, size=9)
    ws.cell(row=3, column=14, value=vehicle.registration_number).alignment = center_align
    ws.cell(row=4, column=13, value="TOTAL AMOUNT").font = Font(bold=True, size=9)
    ws.cell(row=4, column=14, value="VEHICLE DELIVERY DATE").font = Font(bold=True, size=9)
    ws.cell(row=4, column=15).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    ws.cell(row=5, column=13, value="EMI DATE").font = Font(bold=True, size=9)
    ws.cell(row=5, column=14, value=datetime.now().strftime('%d/%m/%Y'))

    # Headers (row 6)
    headers = [
        'DATE OUT', 'TIME OUT', 'DATE IN', 'TIME IN', 'CUSTOMER', 'CONTACT NO.',
        'CUSTOMER ID', 'C/O', 'DESTINATION', 'DAYS OF RENT', 'RENT/DAY',
        'ADV. AMOUNT', 'STARTING KM', 'ENDING KM', 'TOTAL AMOUNT RECEIVED', 'BALANCE'
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_thin

    # Set column widths
    column_widths = [12, 10, 12, 10, 18, 14, 14, 10, 16, 12, 10, 12, 12, 12, 18, 12]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width

    # Data rows
    row_num = 7
    total_received = 0
    for rental in rentals:
        ws.cell(row=row_num, column=1, value=rental.date_out.strftime('%d-%m-%y')).border = border_thin
        ws.cell(row=row_num, column=1).alignment = center_align

        ws.cell(row=row_num, column=2, value=rental.time_out.strftime('%I:%M %p') if rental.time_out else '').border = border_thin
        ws.cell(row=row_num, column=2).alignment = center_align

        ws.cell(row=row_num, column=3, value=rental.date_in.strftime('%d-%m-%y') if rental.date_in else '').border = border_thin
        ws.cell(row=row_num, column=3).alignment = center_align

        ws.cell(row=row_num, column=4, value=rental.time_in.strftime('%I:%M %p') if rental.time_in else '').border = border_thin
        ws.cell(row=row_num, column=4).alignment = center_align

        ws.cell(row=row_num, column=5, value=rental.customer_name).border = border_thin
        ws.cell(row=row_num, column=6, value=rental.contact_no or '').border = border_thin
        ws.cell(row=row_num, column=6).alignment = center_align

        ws.cell(row=row_num, column=7, value=rental.customer_id or '').border = border_thin
        ws.cell(row=row_num, column=8, value=rental.care_of or '').border = border_thin
        ws.cell(row=row_num, column=9, value=rental.destination or '').border = border_thin

        cell_days = ws.cell(row=row_num, column=10, value=float(rental.days_of_rent))
        cell_days.border = border_thin
        cell_days.alignment = center_align

        cell_rent = ws.cell(row=row_num, column=11, value=float(rental.rent_per_day))
        cell_rent.border = border_thin
        cell_rent.alignment = right_align

        cell_adv = ws.cell(row=row_num, column=12, value=float(rental.advance_amount))
        cell_adv.border = border_thin
        cell_adv.alignment = right_align

        ws.cell(row=row_num, column=13, value=rental.starting_km or '').border = border_thin
        ws.cell(row=row_num, column=13).alignment = center_align

        ws.cell(row=row_num, column=14, value=rental.ending_km or '').border = border_thin
        ws.cell(row=row_num, column=14).alignment = center_align

        cell_total = ws.cell(row=row_num, column=15, value=float(rental.total_amount_received))
        cell_total.border = border_thin
        cell_total.alignment = right_align

        # Calculate balance
        total_rent = float(rental.days_of_rent) * float(rental.rent_per_day)
        balance = total_rent - float(rental.total_amount_received)
        cell_balance = ws.cell(row=row_num, column=16, value=balance)
        cell_balance.border = border_thin
        cell_balance.alignment = right_align

        total_received += float(rental.total_amount_received)
        row_num += 1

    # Total row
    if rentals:
        ws.cell(row=row_num, column=14, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=row_num, column=14).alignment = right_align
        cell_total = ws.cell(row=row_num, column=15, value=total_received)
        cell_total.font = Font(bold=True)
        cell_total.alignment = right_align
        cell_total.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # ===== EXPENSE SHEET =====
    ws_expense = wb.create_sheet("Expense History")

    # Title
    ws_expense.merge_cells('A1:F1')
    title_cell = ws_expense.cell(row=1, column=1, value=f"{vehicle.name.upper()} - EXPENSES")
    title_cell.font = title_font
    title_cell.alignment = center_align

    # Headers
    expense_headers = ['DATE', 'PARTICULARS', 'PLACE', 'C/O', 'AMOUNT', 'REMARKS']
    for col_idx, header in enumerate(expense_headers, 1):
        cell = ws_expense.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border_thin

    # Set column widths
    ws_expense.column_dimensions['A'].width = 12
    ws_expense.column_dimensions['B'].width = 30
    ws_expense.column_dimensions['C'].width = 20
    ws_expense.column_dimensions['D'].width = 15
    ws_expense.column_dimensions['E'].width = 15
    ws_expense.column_dimensions['F'].width = 20

    # Data rows
    row_num = 4
    total_expense = 0
    for expense in expenses:
        ws_expense.cell(row=row_num, column=1, value=expense.date.strftime('%d-%m-%y')).border = border_thin
        ws_expense.cell(row=row_num, column=1).alignment = center_align

        ws_expense.cell(row=row_num, column=2, value=expense.particulars).border = border_thin
        ws_expense.cell(row=row_num, column=3, value=expense.place or '').border = border_thin
        ws_expense.cell(row=row_num, column=4, value=expense.care_of or '').border = border_thin

        cell_amount = ws_expense.cell(row=row_num, column=5, value=float(expense.amount))
        cell_amount.border = border_thin
        cell_amount.alignment = right_align

        ws_expense.cell(row=row_num, column=6, value='').border = border_thin

        total_expense += float(expense.amount)
        row_num += 1

    # Total row
    if expenses:
        ws_expense.cell(row=row_num, column=4, value="TOTAL:").font = Font(bold=True)
        ws_expense.cell(row=row_num, column=4).alignment = right_align
        cell_total = ws_expense.cell(row=row_num, column=5, value=total_expense)
        cell_total.font = Font(bold=True)
        cell_total.alignment = right_align
        cell_total.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

